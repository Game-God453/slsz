import json
import os

from django.http import FileResponse, HttpResponse
from django.core.files.storage import default_storage
from django.utils import timezone
from django.views.decorators.http import require_http_methods
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment
from openpyxl.utils import get_column_letter
from datetime import datetime
from competition_space.models import CompetitionSpace, SpaceUser
from files.models import CompetitionFile
from teams.models import Team
from utils import decrypt, custom_response, encrypt


@require_http_methods(["POST"])
def upload_competition_file(request):
    try:
        """上传文件到竞赛空间"""
        space_id = int(decrypt(request.POST.get("space_id")))
        space = CompetitionSpace.objects.get(id=space_id)

        # 检查用户是否有权限上传文件
        if not SpaceUser.objects.get(space=space, user=request.user0).is_admin:
            return custom_response(message="没有权限", status=403)

        # 获取上传的文件
        file = request.FILES.get('file')
        if not file:
            return custom_response(message="没有上传的文件", status=403)

        description = request.POST.get('description')

        # 创建CompetitionFile对象
        competition_file = CompetitionFile.objects.create(
            space=space,
            file=file,
            description=description,
            uploaded_by=request.user0
        )

        return custom_response(message="文件上传成功",status=200)

    except CompetitionSpace.DoesNotExist:
        return custom_response(message="竞赛空间不存在",status=404)
    except Exception as e:
        return custom_response(message=str(e), status=500)

def get_competition_file(request,space_id):
    try:
        space = CompetitionSpace.objects.get(id=int(decrypt(space_id)))
        if not SpaceUser.objects.filter(space=space, user=request.user0).exists():
            return custom_response(message="不是竞赛空间用户，没有文件获取权限", status=403)

        files = CompetitionFile.objects.filter(space=space).order_by('-uploaded_at')
        data = [
            {
                "file_id": encrypt(file.id),
                "description": file.description,
                "uploaded_at": file.uploaded_at.strftime("%Y-%m-%d %H:%M:%S"),
            }
            for file in files
        ]

        return custom_response(data=data, message="获取文件成功", status=200)

    except CompetitionSpace.DoesNotExist:
        return custom_response(message="竞赛空间不存在", status=404)
    except Exception as e:
        return custom_response(message=str(e), status=500)


@require_http_methods(["GET"])
def download_competition_file(request, file_id):
    """下载指定的文件"""
    try:
        file_id = int(decrypt(file_id))
        competition_file = CompetitionFile.objects.get(id=file_id)

        # 获取文件路径
        file_path = competition_file.file.path

        if not SpaceUser.objects.filter(space=competition_file.space, user=request.user0).exists():
            return custom_response(message="不是竞赛空间用户，没有文件获取权限", status=403)

        # 检查文件是否存在
        if not os.path.exists(file_path):
            raise CompetitionFile.DoesNotExist

        # 返回文件下载响应
        response = FileResponse(open(file_path, 'rb'))
        response['Content-Disposition'] = f'attachment; filename="{os.path.basename(file_path)}"'
        response['message'] = '获取文件成功'
        response['status'] = 200  # 自定义头，表示成功状态
        return response
    except CompetitionFile.DoesNotExist:
        return custom_response(message="该文件不存在", status=404)
    except Exception as e:
        return custom_response(message=str(e), status=500)


@require_http_methods(["POST"])
def export_locked_teams_report(request, space_id):
    """导出已锁定的队伍信息到Excel"""
    try:
        space = CompetitionSpace.objects.get(id=int(decrypt(space_id)))

        if not SpaceUser.objects.get(space=space, user=request.user0).is_admin:
            return custom_response(message="没有权限", status=403)

        # 获取前端传来的筛选条件
        data = json.loads(request.body)
        include_fields = {
            'team_name': data.get('include_team_name', True),
            'real_name': data.get('include_real_name', True),
            'gender': data.get('include_gender', True),
            'student_id': data.get('include_student_id', True),
            'college_name': data.get('include_college_name', True),
            'phone_number': data.get('include_phone_number', True),
            'email': data.get('include_email', True),
        }

        # 获取所有已锁定的队伍
        locked_teams = Team.objects.filter(space=space, is_locked=True)

        if not locked_teams.exists():
            return custom_response(message="暂无已锁定的队伍", status=404)

        # 创建Excel工作簿
        wb = Workbook()
        ws = wb.active
        ws.title = "队伍信息"

        # 计算最大队伍人数
        max_team_size = max(team.members.count() for team in locked_teams)

        # 构建表头
        headers = []
        if include_fields['team_name']:
            headers.append("队伍名称")

        for i in range(max_team_size):
            if include_fields['real_name']:
                headers.append(f"队员{i}姓名")
            if include_fields['gender']:
                headers.append(f"队员{i}性别")
            if include_fields['student_id']:
                headers.append(f"队员{i}学号")
            if include_fields['college_name']:
                headers.append(f"队员{i}学院")
            if include_fields['phone_number']:
                headers.append(f"队员{i}电话")
            if include_fields['email']:
                headers.append(f"队员{i}邮箱")

        # 写入表头
        for col_num, header in enumerate(headers, 1):
            col_letter = get_column_letter(col_num)
            ws[f'{col_letter}1'] = header
            ws[f'{col_letter}1'].font = Font(bold=True)
            ws[f'{col_letter}1'].alignment = Alignment(horizontal='center')

        # 写入队伍数据
        row_num = 2
        for team in locked_teams:
            members = team.members.all()
            captain = SpaceUser.objects.filter(space=team.space, team=team, is_captain=True).first()
            members_list = list(members.exclude(id=captain.id)) if captain else list(members)

            col_num = 1
            if include_fields['team_name']:
                ws.cell(row=row_num, column=col_num, value=team.teamName)
                col_num += 1

            for member in [captain] + members_list:
                if member:
                    if include_fields['real_name']:
                        ws.cell(row=row_num, column=col_num, value=member.realName)
                        col_num += 1
                    if include_fields['gender']:
                        ws.cell(row=row_num, column=col_num, value=member.gender)
                        col_num += 1
                    if include_fields['student_id']:
                        ws.cell(row=row_num, column=col_num, value=member.studentId)
                        col_num += 1
                    if include_fields['college_name']:
                        ws.cell(row=row_num, column=col_num, value=member.collegeName)
                        col_num += 1
                    if include_fields['phone_number']:
                        ws.cell(row=row_num, column=col_num, value=member.phoneNumber)
                        col_num += 1
                    if include_fields['email']:
                        ws.cell(row=row_num, column=col_num, value=member.email)
                        col_num += 1

            row_num += 1

        # 将Excel文件内容直接返回给前端
        response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
        response[
            'Content-Disposition'] = f'attachment; filename="{space.title}_teamsReport_{timezone.now().strftime("%Y%m%d_%H%M%S")}.xlsx"'
        wb.save(response)

        return response

    except Exception as e:
        return custom_response(message=str(e), status=500)